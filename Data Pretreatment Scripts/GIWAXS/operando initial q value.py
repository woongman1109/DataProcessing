import os
import math
import pandas as pd
from openpyxl import Workbook

def read_excel_files_n_column(folder_path, n):
    # 폴더 내 모든 엑셀 파일 찾기
    excel_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx') or f.endswith('.xls')]

    # 결과를 저장할 배열
    result_array = []

    # 각 엑셀 파일 읽기
    for file in excel_files:
        file_path = os.path.join(folder_path, file)
        try:
            # 엑셀 파일의 첫 번째 시트를 읽음
            df = pd.read_excel(file_path, header=None)

            # n열 값이 있는지 확인하고, 값이 있으면 배열에 추가
            if df.shape[1] > n:
                column_values = df.iloc[:, n].tolist()  # n열 값을 리스트로 변환
            else:
                column_values = []  # n열이 없으면 빈 리스트로 처리

            # 파일 이름과 n열의 값을 배열에 추가
            result_array.append((file, column_values))

        except Exception as e:
            print(f"Error reading {file}: {e}")
            result_array.append((file, []))
    
    return result_array

def save_to_excel(output_file, data):
    # 새로운 워크북 생성
    wb = Workbook()
    QSheet = wb.create_sheet("q value")
    dSheet = wb.create_sheet("d-Spacing")

    # d-spacing으로 변환
    dSpacing = []
    for c in data:
        dCalc = []
        for q in c[1]:
            dCalc.append(2*math.pi/q)
        dSpacing.append((c[0],dCalc))
        
    # q 데이터 추가
    for row in data:
        # 파일 이름과 n열 값을 엑셀에 각각 추가
        file_name, column_values = row
        QSheet.append([file_name] + column_values)
    
    # d-Spacing 데이터 추가
    for row in dSpacing:
        file_name, column_values = row
        dSheet.append([file_name] + column_values)

    # 엑셀 파일 저장
    wb.save(output_file)
    print(f"Data successfully saved to {output_file}")

# 사용 예시
folder_path = 'C:/Users/woong/Taewoong/SKKU - SAINT/SAINT-OSL/1. Research tmp/2024_Operando GIWAXS/Experiments/GIWAXS/240906_Operando GIWAXS/fitted data'  # 엑셀 파일이 있는 폴더 경로
n = 9  # 예: 세 번째 열 (n은 0부터 시작)
result = read_excel_files_n_column(folder_path, n)

# 결과를 엑셀 파일로 저장
output_file = 'test.xlsx'  # 저장할 파일 이름
save_to_excel(output_file, result)
