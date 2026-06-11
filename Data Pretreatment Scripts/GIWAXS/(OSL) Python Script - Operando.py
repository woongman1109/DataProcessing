############################################
## Organic Semiconductor Lab, SAINT, SKKU ##
## Source code copyright to Taewoong Yoon ##
## +++++ Email: TWYoon.rs@gmail.com +++++ ##
## ░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░ ##
## ░░░░░░░█████░░░░███████░░██░░░░░░░░░░░ ##
## ░░░░░██░░░░░██░██░░░░░██░██░░░░░░░░░░░ ##
## ░░░░░█░░░░░░░█░███░░░░░░░██░░░░░░░░░░░ ##
## ░░░░░█░░░░░░░█░░░█████░░░██░░░░░░░░░░░ ##
## ░░░░░█░░░░░░░█░░░░░░░███░██░░░░░░░░░░░ ##
## ░░░░░██░░░░░██░██░░░░░██░██░░░░░██░░░░ ##
## ░░░░░░░█████░░░░███████░░█████████░░░░ ##
## ░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░ ##
## ++++++++++++++++++++++++++++++++++++++ ##
############################################

## ++++++++++++ GIWAXS DATA from 3C, PAL ++++++++++++ ##
## ---------------- Gather all data into one .xlsx file.
## ----------------------------- File name organization.

import os
import openpyxl as xl

############### Functions ################
def sheetName(thisData, _dataIndex, isExist=''):
    _wsname = ''
    _default = ''
    i=0
 
    for s in os.path.splitext(thisData)[0]:
        if i<=20 and i <= len(thisData):
            _default += s
        i+=1
    if '_90degree' in n:
        _default+='..._Qz'
    elif '_0degree' in n:
        _default+='..._Qxy'
    
    if isExist == '':
        _newname = input('|{0:^7}| {1:<59} | {2:<30} | '.format(_dataIndex, thisData, _default))
    else:
        _newname = input('|{0:^7}| {1:<59} | {2:<30} | '.format(_dataIndex, thisData, isExist))

    while len(_newname) >= 30:
        _newname = input('|       |   - Sheet name should be less than 31 characters.. (Now: {0:>3})                  Rename this:  | '.format(len(_newname)))
        

    if _newname == '':
        _wsname = _default
    else:
        _wsname = _newname
    
    # print(_wsname)
    return _wsname

################ Start ################

print("Organic Semiconductor Lab, SAINT, SKKU", "Source code copyright to Taewoong Yoon", "+++++ Email: TWYoon.rs@gmail.com +++++", "░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░", "░░░░░░░█████░░░░███████░░██░░░░░░░░░░░", "░░░░░██░░░░░██░██░░░░░██░██░░░░░░░░░░░", "░░░░░█░░░░░░░█░███░░░░░░░██░░░░░░░░░░░", "░░░░░█░░░░░░░█░░░█████░░░██░░░░░░░░░░░", "░░░░░█░░░░░░░█░░░░░░░███░██░░░░░░░░░░░", "░░░░░██░░░░░██░██░░░░░██░██░░░░░██░░░░", "░░░░░░░█████░░░░███████░░█████████░░░░", "░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░", "++++++++++++++++++++++++++++++++++++++", sep="\n")
print('\n\nNow Running: ', os.path.realpath(__file__), '\n')
fileList = []
for file in os.listdir():
    if '.csv' in file:
        fileList.append(file)

print(fileList)

print('============ Name of your file (EXCEPT EXTENSION) ============')
outname = input("Filename (except .xlsx): ")
print('==============================================================')

out = xl.Workbook()
ws1 = out.worksheets[0]
ws1.title = "DataList"
ws1.cell(1,1,'FileName')
ws1.cell(1,2,"SheetName")

### Declare variables ###
oneDplotList = []
wsname = ''

dataIndex = 1
print('     1. SheetName should be shorter than 30 characters')
print('     2. Press Enter to use Default name                                                                           1         2         3')
print('+--------------------------------------------------------------------------------------------------------1--------0---------0---------0--+')
print('|{0:^7}| {1:<59} | {2:<30} | {3:<32}|'.format('index', ws1[1][0].value, 'Default', 'SheetName (less than 30)'))
for fileName in fileList:
    ## Set Sheetname ##
    wsname = sheetName(n, dataIndex)
    data = out.create_sheet(wsname)
    oneDplotList.append(n)
    # print(os.path.splitext(fileName)[0] + '\\' + n)

    ## Write file ##
    rawcsv = open(os.path.splitext(fileName)[0] + '\\' + n, 'r', encoding='UTF8')
    l = 1
    for line in rawcsv:
        if float(line.split(',')[0]) >= 0 and float(line.split(',')[1][:-1]) != 0:
            data.cell(l, 1, line.split(',')[0])
            data.cell(l, 2, line.split(',')[1][:-1])
            l+=1
    rawcsv.close()

    ## Update DataList Sheet ##
    ws1.cell(len(oneDplotList)+1,1,n)
    ws1.cell(len(oneDplotList)+1,2,wsname)
    dataIndex += 1
print('+-------------------------------------------------------------------------------------------------------+')

# print('\n+{0:-^16} Data List {0:-^15}+'.format('-'))
# print(oneDplotList)

_isReset = True
while _isReset == True:
    if input("Is there anything to change? (y/n)") == 'y':
        resetList = input("Index of the file you want to change (seperate with ','): ").split(',')
        print('\n     1. SheetName should be shorter than 30 characters')
        print('     2. Press Enter to use Default name                                                                           1         2         3')
        print('+--------------------------------------------------------------------------------------------------------1--------0---------0---------0--+')
        print('|{0:^7}| {1:<59} | {2:<30} | {3:<32}|'.format('index', ws1[1][0].value, 'Current SheetName', 'New SheetName (less than 30)'))
        for i in resetList:
            out.worksheets[int(i)].title = sheetName(ws1[int(i)+1][0].value,int(i),out.worksheets[int(i)].title)
            ws1.cell(int(i)+1,2,out.worksheets[int(i)].title)
    else:
        _isReset = False
        print(_isReset)
        continue



print('\nCheck the result below')
print('+-------------------------------------------------------------------------------------------------------+')
rni = 1
print('|{0:^7}| {1:<59}| {2:<33}|'.format('index', ws1[1][0].value, 'SheetName'))
for rn in ws1.rows:
    if rni == 1:
        rni+=1
        continue
    else:
        print('|{0:^7}| {1:<59}| {2:<33}|'.format(rni-1, rn[0].value, rn[1].value))
        rni += 1
print('+-------------------------------------------------------------------------------------------------------+')

out.save(outname+'.xlsx')
input('Press Enter to exit')
