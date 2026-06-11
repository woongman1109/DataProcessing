import openpyxl as xl

chemical_sheet = xl.load_workbook("chemicals.xlsx", data_only=True)

general = chemical_sheet['general']
OSC = chemical_sheet['OSC']

RAW_general_old = general['A2':'A300']
RAW_general_new = general['B2':'B300']
RAW_OSC_old = OSC['A2':'A300']
RAW_OSC_new = OSC['B2':'B300']

general_old = []
general_new = []
OSC_old = []
OSC_new = []

###############################
for row in RAW_general_old:
    for cell in row:
        if cell.value != 'None':
            general_old.append(cell.value)

for row in RAW_general_new:
    for cell in row:
        if cell.value != 'None':
            general_new.append(cell.value)

for row in RAW_OSC_old:
    for cell in row:
        if cell.value != 'None':
            OSC_old.append(cell.value)

for row in RAW_OSC_new:
    for cell in row:
        if cell.value != 'None':
            OSC_new.append(cell.value)

################################
checkfile = xl.Workbook()
checked_general = checkfile.create_sheet('general')
checked_OSC = checkfile.create_sheet('OSC')

onlyInGeneralOld = []
onlyInGeneralNew = []
onlyInOSCOld = []
onlyInOSCNew = []

i=1
for ch in general_old:
    if ch in general_old and ch not in general_new:
        onlyInGeneralOld.append(ch)
        checked_general.cell(i,1,ch)
        i+=1
        # print("Appending to OLDs: ", ch)

i=1
for ch in general_new:
    if ch in general_new and ch not in general_old:
        onlyInGeneralNew.append(ch)
        checked_general.cell(i,2,ch)
        i+=1
        # print("Appending to NEWs: ", ch)

i=1
for ch in OSC_old:
    if ch in OSC_old and ch not in OSC_new:
        onlyInOSCOld.append(ch)
        checked_OSC.cell(i,1,ch)
        i+=1
        # print("Appending to OLDs: ", ch)

i=1
for ch in OSC_new:
    if ch in OSC_new and ch not in OSC_old:
        onlyInOSCNew.append(ch)
        checked_OSC.cell(i,2,ch)
        i+=1
        # print("Appending to NEWs: ", ch)

checkfile.save('checkedfile.xlsx')